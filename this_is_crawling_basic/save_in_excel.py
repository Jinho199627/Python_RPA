import openpyxl

wb = openpyxl.Workbook() # 엑셀 파일 만들기
ws = wb.create_sheet('첫번째_시트') # 엑셀 워크시트 만들기

ws['A1'] = '부서' # 데이터 추가하기
ws['B1'] = '이름'
ws['A2'] = 'IT' # 데이터 추가하기
ws['B2'] = 'Jinho'

wb.save(r'C:\Users\jinho\Desktop\list.xlsx') # 경로 지정 안하면 현재 작업 폴더에 저장된다.
# 앞에 r을 붙여 원래 문자 형태로 만들기