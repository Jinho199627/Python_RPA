import openpyxl

fpath = r'C:\Users\jinho\Desktop\list.xlsx'

wb = openpyxl.load_workbook(fpath) # 파일 열기
ws = wb['첫번째_시트'] # 엑셀 시트 선택

ws['A3'] = 'Sales' # 데이터 수정하기
ws['B3'] = 'James'

wb.save(fpath) # 파일 저장